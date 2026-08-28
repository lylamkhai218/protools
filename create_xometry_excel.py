import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_xometry_competencies_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xometry Competencies Survey"
    ws.views.sheetView[0].showGridLines = True

    # Palette
    NAVY_DARK = "003366"
    BLUE_PRIMARY = "00478D"
    BLUE_LIGHT = "EBF2FA"
    GRAY_HEADER = "E2E8F0"
    GRAY_LIGHT = "F8FAFC"
    BORDER_COLOR = "CBD5E1"
    ACCENT_GREEN = "15803D"
    ACCENT_AMBER = "B45309"

    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="E2E8F0")
    font_kpi_label = Font(name="Segoe UI", size=9, bold=True, color="475569")
    font_kpi_val = Font(name="Segoe UI", size=14, bold=True, color="00478D")
    font_cat_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=9.5)
    font_bold = Font(name="Segoe UI", size=9.5, bold=True)
    font_en = Font(name="Segoe UI", size=9.5, bold=True, color="0F172A")
    font_vn = Font(name="Segoe UI", size=9.5, color="334155")

    fill_title = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_header = PatternFill(start_color=BLUE_PRIMARY, end_color=BLUE_PRIMARY, fill_type="solid")
    fill_kpi = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid")
    fill_zebra = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    cat_fills = {
        "CNC": PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"),
        "SHEET": PatternFill(start_color="0369A1", end_color="0369A1", fill_type="solid"),
        "CASTING": PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid"),
        "3D": PatternFill(start_color="7E22CE", end_color="7E22CE", fill_type="solid"),
        "OTHER": PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid"),
        "POST": PatternFill(start_color="475569", end_color="475569", fill_type="solid"),
    }

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    thick_bottom = Border(
        bottom=Side(style="medium", color=NAVY_DARK),
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
    )

    # 1. Header Banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "BẢNG KHẢO SÁT NĂNG LỰC SẢN XUẤT (XOMETRY PARTNER ONBOARDING)"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = "Dành cho Giám đốc Sản xuất / Trưởng phòng Kỹ thuật rà soát & tích chọn năng lực thực tế để khớp đơn hàng Xometry Partner Network"
    ws["A2"].font = font_subtitle
    ws["A2"].fill = fill_title
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 2. KPI Summary Boxes
    ws.merge_cells("A4:B4")
    ws["A4"] = "TỔNG NĂNG LỰC KHẢO SÁT"
    ws["A4"].font = font_kpi_label
    ws["A4"].fill = fill_kpi
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A5:B5")
    ws["A5"] = '=COUNTA(C8:C70)-6'
    ws["A5"].font = font_kpi_val
    ws["A5"].fill = fill_kpi
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("C4:D4")
    ws["C4"] = "SỐ NĂNG LỰC CÔNG TY ĐẠT (YES)"
    ws["C4"].font = font_kpi_label
    ws["C4"].fill = fill_kpi
    ws["C4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("C5:D5")
    ws["C5"] = '=COUNTIF(E8:E70, "*ĐẠT*")'
    ws["C5"].font = Font(name="Segoe UI", size=14, bold=True, color=ACCENT_GREEN)
    ws["C5"].fill = fill_kpi
    ws["C5"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("E4:F4")
    ws["E4"] = "NĂNG LỰC ĐANG ĐẦU TƯ / OUTSOURCE"
    ws["E4"].font = font_kpi_label
    ws["E4"].fill = fill_kpi
    ws["E4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("E5:F5")
    ws["E5"] = '=COUNTIF(E8:E70, "*ĐẦU TƯ*") + COUNTIF(E8:E70, "*LIÊN KẾT*")'
    ws["E5"].font = Font(name="Segoe UI", size=14, bold=True, color=ACCENT_AMBER)
    ws["E5"].fill = fill_kpi
    ws["E5"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G4:H4")
    ws["G4"] = "NGƯỜI ĐÁNH GIÁ & NGÀY RÀ SOÁT"
    ws["G4"].font = font_kpi_label
    ws["G4"].fill = fill_kpi
    ws["G4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G5:H5")
    ws["G5"] = "GĐ Sản xuất: ........................ | Ngày: ..../..../2026"
    ws["G5"].font = font_data
    ws["G5"].fill = fill_kpi
    ws["G5"].alignment = Alignment(horizontal="center", vertical="center")

    for row in range(4, 6):
        ws.row_dimensions[row].height = 22
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border

    # 3. Table Column Headers
    headers = [
        ("STT", 6),
        ("Nhóm công nghệ\n(Category)", 22),
        ("Phương pháp sản xuất\n(Production Method - EN)", 30),
        ("Dịch nghĩa kỹ thuật & Diễn giải\n(Technical Translation - VN)", 42),
        ("LỰA CHỌN\n(Selection)", 24),
        ("Quy mô máy móc / Công suất hiện có\n(Machine Model / Specs / Bed Size)", 38),
        ("Dung sai / Tiêu chuẩn đạt được\n(Tolerance / ISO Standards)", 28),
        ("Ghi chú của GĐ Sản xuất\n(Production Notes)", 30)
    ]

    header_row = 7
    ws.row_dimensions[header_row].height = 32
    for col_idx, (header_text, width) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thick_bottom
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data Taxonomy from Screenshots
    data = [
        # GROUP 1: CNC
        ("CAT", "1. CNC MACHINING (GIA CÔNG CƠ KHÍ CHÍNH XÁC CNC)", "CNC"),
        ("Milling", "Phay CNC 3 trục / 4 trục (Vertical / Horizontal Machining Center)", "CNC"),
        ("Milling 5 axes", "Phay CNC 5 trục đồng thời / 3+2 trục chính xác cao", "CNC"),
        ("Milling mass", "Phay sản xuất hàng loạt / khối lượng lớn (Mass Production)", "CNC"),
        ("Turning", "Tiện CNC tiêu chuẩn (2 trục X, Z)", "CNC"),
        ("Turning 2+ axes", "Tiện CNC đa trục (Tiện phay kết hợp Live Tooling / C-axis, Y-axis)", "CNC"),
        ("Turning Automat", "Tiện đùn tự động kiểu Thụy Sĩ (Swiss Lathe / Auto Screw Machining)", "CNC"),
        ("Grinding", "Mài chính xác (Mài phẳng Surface, mài tròn Cylindrical, mài vô tâm)", "CNC"),
        ("Polishing", "Đánh bóng cơ học / Hoàn thiện bóng gương bề mặt chi tiết", "CNC"),
        ("Slotting", "Xọc rãnh then / Cắt rãnh kỹ thuật (Slotting / Broaching)", "CNC"),
        ("Liner Honing", "Doa chính xác / Đánh bóng honing lòng xi lanh, ống lót", "CNC"),
        ("Thread Rolling", "Cán ren / Lăn ren áp lực cao (Độ bền ren vượt trội)", "CNC"),

        # GROUP 2: Sheet Metal
        ("CAT", "2. SHEET METAL & FABRICATION (GIA CÔNG & ĐỊNH HÌNH KIM LOẠI TẤM)", "SHEET"),
        ("Laser Cutting", "Cắt Laser Fiber kim loại tấm (Thép, Inox, Nhôm, Đồng)", "SHEET"),
        ("Waterjet Cutting", "Cắt tia nước có hạt mài (Cắt kim loại dày, đá, composite, không sinh nhiệt)", "SHEET"),
        ("Plasma/Gas Cutting", "Cắt Plasma CNC / Cắt Oxy-Gas bản mã thép tấm dày", "SHEET"),
        ("Bending Sheet", "Chấn / Gập tôn tấm CNC (Press Brake)", "SHEET"),
        ("Bending Tube", "Uốn ống kim loại định hình (Tube Bending)", "SHEET"),
        ("Bending Wire", "Uốn dây kim loại / Dây thép định hình CNC (Wire Bending)", "SHEET"),
        ("Bending bar", "Uốn thanh / Thép hình / Thép cây (Bar Bending)", "SHEET"),
        ("Punching", "Đột dập CNC kim loại tấm (CNC Turret Punching)", "SHEET"),
        ("Stamping", "Dập định hình liên hoàn / Dập khuôn khối (Progressive / Deep Draw Stamping)", "SHEET"),

        # GROUP 3: Casting & Molding
        ("CAT", "3. CASTING, MOLDING & FORMING (ĐÚC, ÉP PHUN & TẠO HÌNH)", "CASTING"),
        ("Sand molded casting", "Đúc khuôn cát (Sand Casting) - Gang, Thép, Hợp kim", "CASTING"),
        ("Metal Injection Molding", "Ép phun kim loại (MIM) - Chi tiết nhỏ, độ phức tạp cao", "CASTING"),
        ("Compression Molding", "Ép nén định hình (Cao su, composite, nhựa nhiệt rắn)", "CASTING"),
        ("Investment Casting", "Đúc mẫu chảy chính xác (Đúc sáp / Mất sáp Lost-wax Casting)", "CASTING"),
        ("Ceramic Injection Molding", "Ép phun gốm kỹ thuật cao (CIM)", "CASTING"),
        ("LSR Injection Molding", "Ép phun cao su Silicone lỏng (Liquid Silicone Rubber)", "CASTING"),
        ("Die cutting", "Bế cắt khuôn định hình (Gioăng, seal, màng nhựa, đệm xốp)", "CASTING"),
        ("Extrusion", "Đùn ép định hình (Đùn nhôm định hình, đùn nhựa profile)", "CASTING"),
        ("Pressure Forming", "Tạo hình áp lực cao kim loại / tấm", "CASTING"),
        ("Twin Sheet Forming", "Định hình nhiệt hai tấm rỗng chân không (Twin Sheet Thermoforming)", "CASTING"),
        ("Injection Molding", "Ép phun nhựa nhiệt dẻo tiêu chuẩn (Plastic Injection Molding)", "CASTING"),
        ("In-Mold Decoration", "Trang trí bề mặt trong khuôn đúc (IMD)", "CASTING"),
        ("In-Mold Filming", "Dán màng hoa văn kỹ thuật trong khuôn (IMF)", "CASTING"),
        ("Die Casting", "Đúc áp lực cao buồng nóng/buồng lạnh (Nhôm, Kẽm, Magie)", "CASTING"),
        ("Vacuum Casting", "Đúc hút chân không khuôn silicon (Tạo mẫu nhanh nhựa PU/ABS)", "CASTING"),
        ("Thermoforming", "Định hình nhiệt chân không tấm nhựa (Thermoforming)", "CASTING"),
        ("Blow Molding", "Ép thổi rỗng chai, can, thùng chứa (Blow Molding)", "CASTING"),
        ("Foam Molding", "Đúc tạo hình bọt xốp / Mút định hình (Foam Molding)", "CASTING"),
        ("Vacuum Forming", "Hút nổi định hình chân không màng mỏng (Vacuum Forming)", "CASTING"),

        # GROUP 4: 3D Printing
        ("CAT", "4. 3D PRINTING & ADDITIVE MANUFACTURING (IN 3D & BỒI ĐẮP)", "3D"),
        ("3D Printing (SLS)", "In 3D thiêu kết laser chọn lọc bột nhựa (Selective Laser Sintering - PA12/Nylon)", "3D"),
        ("3D Printing (MJF)", "In 3D công nghệ Multi Jet Fusion (HP MJF - Nhựa kỹ thuật chịu lực)", "3D"),
        ("3D Printing (SLA)", "In 3D quang trùng hợp nhựa lỏng độ mịn cao (Stereolithography - Resin)", "3D"),
        ("3D Printing (FDM)", "In 3D đùn sợi nhựa nhiệt dẻo (Fused Deposition Modeling - PLA/ABS/PETG/PEEK)", "3D"),
        ("3D Printing (DLS)", "In 3D ánh sáng kỹ thuật số siêu tốc (Digital Light Synthesis - Carbon DLS)", "3D"),
        ("3D Printing (DMLS)", "In 3D thiêu kết laser kim loại trực tiếp (Direct Metal Laser Sintering - Inox, Titan, Nhôm)", "3D"),
        ("3D Printing (Polyjet)", "In 3D phun hạt photopolymer đa vật liệu / đa màu sắc (PolyJet)", "3D"),

        # GROUP 5: Other Methods
        ("CAT", "5. OTHER METHODS & FABRICATION (PHƯƠNG PHÁP CHẾ TẠO KHÁC)", "OTHER"),
        ("Tooth Gear", "Gia công chế tạo bánh răng (Phay lăn răng Hobbing, xọc răng, mài răng chính xác)", "OTHER"),
        ("EDM", "Gia công tia lửa điện (Cắt dây Wire-EDM & Xung định hình Die-Sinker EDM)", "OTHER"),
        ("Assembly", "Lắp ráp cụm chi tiết cơ khí, đóng gói module (Mechanical Assembly)", "OTHER"),
        ("Welding Steel", "Hàn kết cấu thép (Hàn MIG, MAG, TIG, que hàn tiêu chuẩn)", "OTHER"),
        ("Welding Alu", "Hàn nhôm và hợp kim nhôm chuyên dụng (AC TIG / Pulse MIG)", "OTHER"),
        ("Welding Stainless Steel", "Hàn inox / thép không gỉ thẩm mỹ & vi sinh (Sanitary TIG)", "OTHER"),
        ("Molds&Dies", "Thiết kế & Chế tạo khuôn mẫu chính xác (Khuôn ép nhựa, khuôn dập kim loại)", "OTHER"),
        ("Assembly machining", "Gia công cơ khí tinh chỉnh sau khi đã lắp ghép cụm chi tiết", "OTHER"),

        # GROUP 6: Post Processing
        ("CAT", "6. POST PROCESSING & FINISHING (XỬ LÝ BỀ MẶT & HOÀN THIỆN)", "POST"),
        ("Post Processing", "Xử lý nhiệt luyện (Tôi, thấm Carbon, ram) & Xử lý bề mặt (Anodize, Xi mạ Niken/Kẽm, Sơn tĩnh điện, Bắn cát Bead Blast)", "POST"),
    ]

    # Data Validation List
    dv = DataValidation(
        type="list",
        formula1='"☑ ĐẠT (Có máy tại xưởng), ☐ KHÔNG (Chưa làm), ⏳ ĐANG ĐẦU TƯ, 🤝 LIÊN KẾT GIA CÔNG"',
        allow_blank=True
    )
    ws.add_data_validation(dv)

    current_row = 8
    stt_counter = 1

    for item in data:
        ws.row_dimensions[current_row].height = 24
        if item[0] == "CAT":
            cat_name = item[1]
            cat_key = item[2]
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            cell = ws.cell(row=current_row, column=1, value=cat_name)
            cell.font = font_cat_header
            cell.fill = cat_fills.get(cat_key, fill_header)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for c in range(1, 9):
                ws.cell(row=current_row, column=c).border = thin_border
            ws.row_dimensions[current_row].height = 26
        else:
            method_en, method_vn, cat_key = item
            is_zebra = (stt_counter % 2 == 0)
            row_fill = fill_zebra if is_zebra else fill_white

            # Col 1: STT
            c1 = ws.cell(row=current_row, column=1, value=stt_counter)
            c1.font = font_bold
            c1.alignment = Alignment(horizontal="center", vertical="center")

            # Col 2: Category Short
            c2 = ws.cell(row=current_row, column=2, value=cat_key)
            c2.font = font_bold
            c2.alignment = Alignment(horizontal="center", vertical="center")

            # Col 3: Method EN
            c3 = ws.cell(row=current_row, column=3, value=method_en)
            c3.font = font_en
            c3.alignment = Alignment(horizontal="left", vertical="center")

            # Col 4: Method VN
            c4 = ws.cell(row=current_row, column=4, value=method_vn)
            c4.font = font_vn
            c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Col 5: Selection (Dropdown)
            c5 = ws.cell(row=current_row, column=5, value="☐ KHÔNG (Chưa làm)")
            c5.font = font_bold
            c5.alignment = Alignment(horizontal="center", vertical="center")
            dv.add(c5)

            # Col 6: Machine Specs
            c6 = ws.cell(row=current_row, column=6, value="")
            c6.font = font_data
            c6.alignment = Alignment(horizontal="left", vertical="center")

            # Col 7: Tolerance / Standards
            c7 = ws.cell(row=current_row, column=7, value="")
            c7.font = font_data
            c7.alignment = Alignment(horizontal="left", vertical="center")

            # Col 8: Notes
            c8 = ws.cell(row=current_row, column=8, value="")
            c8.font = font_data
            c8.alignment = Alignment(horizontal="left", vertical="center")

            for col_idx in range(1, 9):
                cell_item = ws.cell(row=current_row, column=col_idx)
                cell_item.fill = row_fill
                cell_item.border = thin_border

            stt_counter += 1
        current_row += 1

    # Freeze Panes (Header stays visible when scrolling)
    ws.freeze_panes = "A8"

    output_path = "d:/T&TVina/protools/Xometry_Production_Competencies_Bilingual.xlsx"
    wb.save(output_path)
    print(f"File created successfully at: {output_path}")

if __name__ == "__main__":
    create_xometry_competencies_sheet()
